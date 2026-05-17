import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

FONT_NAME = "Times New Roman"
FONT_SIZE = 11

# Цены блока 1 (металл)
PRICE_METAL = {0.5: 3400, 0.7: 3700, 0.9: 4200, 1.0: 4500}
PRICE_MANUF = {0.5: 4500, 0.7: 4800, 0.9: 5200, 1.0: 5500}

# Тарифы монтажа оборудования (₸/шт)
INSTALL_RATES = {
    "установка": 120000,
    "радиальный": 80000,
    "канальный прямоугольный": 60000,
    "канальный круглый": 45000,
    "прочее": 30000,
}

# Монтаж воздуховодов (₸/м²)
INSTALL_DUCT_RATE = 3500

# Расходники и транспорт
CONSUMABLES_PCT = 0.05
TRANSPORT_PCT = 0.03


def _thin_border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _cell(ws, row, col, value="", bold=False, align="center", num_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=bold)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = _thin_border()
    if num_format:
        c.number_format = num_format
    return c


def _merge_cell(ws, row, col_start, col_end, value="", bold=False, align="center"):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=value)
    c.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=bold)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = _thin_border()
    return c


def generate_kp(project: dict, output_dir: str) -> str:
    """
    Генерирует Excel КП.
    project — словарь проекта с полями:
        - filename: имя файла спецификации
        - systems: список систем с металлом (из claude_service)
        - supplier_data: {"systems": [...]} из RowenSupplier
    Возвращает путь к сгенерированному файлу.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "КП"

    # Ширина колонок: A-наим, B-ед, C-кол, D-цена металл, E-цена изгот, F-сумма
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16

    row = 1

    # ─── ЗАГОЛОВОК ───────────────────────────────────────────────
    ws.row_dimensions[row].height = 20
    _merge_cell(ws, row, 1, 6,
                f"Коммерческое предложение  |  {project.get('filename', '')}  |  {datetime.now().strftime('%d.%m.%Y')}",
                bold=True)
    row += 1

    # ─── БЛОК 1: МЕТАЛЛ ──────────────────────────────────────────
    ws.row_dimensions[row].height = 18
    _merge_cell(ws, row, 1, 6, "1. Изготовление металлических воздуховодов", bold=True)
    row += 1

    # Заголовки блока 1
    headers1 = ["Наименование", "Ед.", "Кол-во (м²)", "Цена за металл", "Цена за изготовление", "Сумма"]
    for i, h in enumerate(headers1, 1):
        _cell(ws, row, i, h, bold=True)
    row += 1

    # Собираем м² по толщинам из всех систем
    thickness_totals = {}
    systems = project.get("systems", [])
    for sys in systems:
        for item in sys.get("items", []):
            t = round(float(item.get("thickness_mm", 0.5)), 1)
            area = float(item.get("area_m2", 0))
            thickness_totals[t] = thickness_totals.get(t, 0) + area

    block1_rows = []
    block1_start = row
    for thickness in sorted(thickness_totals.keys()):
        area = round(thickness_totals[thickness], 1)
        p_metal = PRICE_METAL.get(thickness, 4000)
        p_manuf = PRICE_MANUF.get(thickness, 5000)
        summa = round(area * (p_metal + p_manuf))
        name = f"Изготовление вент. изделий из оцинк. стали, толщиной {thickness}мм"
        _cell(ws, row, 1, name, align="left")
        _cell(ws, row, 2, "м²")
        _cell(ws, row, 3, area, num_format="#,##0.0")
        _cell(ws, row, 4, p_metal, num_format="# ##0")
        _cell(ws, row, 5, p_manuf, num_format="# ##0")
        _cell(ws, row, 6, f"=C{row}*(D{row}+E{row})", num_format="# ##0")
        block1_rows.append(row)
        row += 1

    # Итого блок 1
    if block1_rows:
        _cell(ws, row, 1, "ИТОГО Блок 1:", bold=True, align="left")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        _cell(ws, row, 6, f"=SUM(F{block1_start}:F{row-1})", bold=True, num_format="# ##0")
    block1_total_row = row
    row += 2

    # ─── БЛОК 2: ОБОРУДОВАНИЕ ────────────────────────────────────
    _merge_cell(ws, row, 1, 6, "2. Оборудование (поставщик)", bold=True)
    row += 1

    headers2 = ["Наименование", "Ед.", "Кол-во", "Цена (₸)", "", "Сумма (₸)"]
    for i, h in enumerate(headers2, 1):
        _cell(ws, row, i, h, bold=True)
    row += 1

    supplier_data = project.get("supplier_data", {})
    supplier_systems = supplier_data.get("systems", [])
    block2_start = row
    block2_total_m2 = 0  # для монтажа

    for sys in supplier_systems:
        sys_name = sys.get("system", "")
        # Заголовок системы
        _merge_cell(ws, row, 1, 6, sys_name, bold=True, align="left")
        row += 1

        for item in sys.get("items", []):
            name = item.get("name", "")
            qty = item.get("qty", 1)
            unit = item.get("unit", "шт")
            price_kzt = round(item.get("price_kzt", 0))
            summa_kzt = round(item.get("summa_kzt", 0))
            _cell(ws, row, 1, name, align="left")
            _cell(ws, row, 2, unit)
            _cell(ws, row, 3, qty)
            _cell(ws, row, 4, price_kzt, num_format="# ##0")
            _cell(ws, row, 5, "")
            _cell(ws, row, 6, f"=C{row}*D{row}", num_format="# ##0")
            row += 1

    # Итого блок 2
    _cell(ws, row, 1, "ИТОГО Блок 2:", bold=True, align="left")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    _cell(ws, row, 6, f"=SUM(F{block2_start}:F{row-1})", bold=True, num_format="# ##0")
    block2_total_row = row
    row += 2

    # ─── БЛОК 3: МОНТАЖ ──────────────────────────────────────────
    _merge_cell(ws, row, 1, 6, "3. Монтажные работы", bold=True)
    row += 1

    headers3 = ["Наименование", "Ед.", "Кол-во", "Цена", "", "Сумма"]
    for i, h in enumerate(headers3, 1):
        _cell(ws, row, i, h, bold=True)
    row += 1

    block3_start = row

    # Монтаж воздуховодов
    total_area = sum(thickness_totals.values())
    total_area_r = round(total_area, 1)
    _cell(ws, row, 1, "Монтаж воздуховодов и фасонных изделий", align="left")
    _cell(ws, row, 2, "м²")
    _cell(ws, row, 3, total_area_r, num_format="#,##0.0")
    _cell(ws, row, 4, INSTALL_DUCT_RATE, num_format="# ##0")
    _cell(ws, row, 5, "")
    _cell(ws, row, 6, f"=C{row}*D{row}", num_format="# ##0")
    row += 1

    # Монтаж оборудования по системам
    for sys in supplier_systems:
        for item in sys.get("items", []):
            name = item.get("name", "").lower()
            qty = item.get("qty", 1)
            if any(k in name for k in ["установка", "агрегат", "вентиляционная"]):
                rate = INSTALL_RATES["установка"]
                label = "Монтаж вентиляционной установки"
            elif "радиальный" in name or "вентилятор" in name:
                rate = INSTALL_RATES["радиальный"]
                label = "Монтаж вентилятора радиального"
            elif "канальный" in name:
                rate = INSTALL_RATES["канальный прямоугольный"]
                label = "Монтаж вентилятора канального"
            else:
                continue
            _cell(ws, row, 1, label, align="left")
            _cell(ws, row, 2, "шт")
            _cell(ws, row, 3, qty)
            _cell(ws, row, 4, rate, num_format="# ##0")
            _cell(ws, row, 5, "")
            _cell(ws, row, 6, f"=C{row}*D{row}", num_format="# ##0")
            row += 1

    # Расходные материалы
    _cell(ws, row, 1, "Расходные материалы (крепёж, саморезы, хомуты, лента)", align="left")
    _cell(ws, row, 2, "компл.")
    _cell(ws, row, 3, 1)
    _cell(ws, row, 4, f"=F{block1_total_row}*{CONSUMABLES_PCT}", num_format="# ##0")
    _cell(ws, row, 5, "")
    _cell(ws, row, 6, f"=D{row}", num_format="# ##0")
    row += 1

    # Транспортные расходы
    _cell(ws, row, 1, "Транспортные расходы", align="left")
    _cell(ws, row, 2, "компл.")
    _cell(ws, row, 3, 1)
    _cell(ws, row, 4, f"=(F{block1_total_row}+F{block2_total_row})*{TRANSPORT_PCT}", num_format="# ##0")
    _cell(ws, row, 5, "")
    _cell(ws, row, 6, f"=D{row}", num_format="# ##0")
    row += 1

    # Итого блок 3
    _cell(ws, row, 1, "ИТОГО Блок 3:", bold=True, align="left")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    _cell(ws, row, 6, f"=SUM(F{block3_start}:F{row-1})", bold=True, num_format="# ##0")
    block3_total_row = row
    row += 2

    # ─── ИТОГО ОБЩЕЕ ─────────────────────────────────────────────
    _merge_cell(ws, row, 1, 5, "ИТОГО по КП:", bold=True)
    _cell(ws, row, 6,
          f"=F{block1_total_row}+F{block2_total_row}+F{block3_total_row}",
          bold=True, num_format="# ##0")

    # Сохраняем
    os.makedirs(output_dir, exist_ok=True)
    project_id = project.get("id", "unknown")
    filename = f"КП_{project_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = os.path.join(output_dir, filename)
    wb.save(path)
    return path
